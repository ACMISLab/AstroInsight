import json
import random


import pandas as pd
from tqdm import tqdm
from PaperAgent.API.LLM_API import call_with_deep_jsonout


def json2xlsx(sample_size = 300000):
    # 定义文件路径
    json_file_path = r'E:\PaperAgent\archive\arxiv-metadata-oai-snapshot.json'
    excel_file_path = r'E:\PaperAgent\archive\output'

    # 读取JSON文件到DataFrame
    data_list = []  # 用于存储所有数据的列表
    count = 0
    total_lines = sum(1 for line in open(json_file_path, 'r', encoding='utf-8'))  # 计算文件总行数

    with open(json_file_path, 'r', encoding='utf-8') as file:
        for line in tqdm(file, total=total_lines, desc="Reading JSON"):
            # 只有当随机数小于sample_size与当前已读取行数的比值时才采样（这不是一个严格的均匀采样，但对于大数据集通常足够好）
            # 或者，为了更严格的均匀采样，可以先读取所有行的索引，然后随机选择sample_size个索引，但这样需要更多的内存
            if random.random() < (sample_size / total_lines) and len(data_list) < sample_size:
                dic = json.loads(line)
                data = {
                    'Authors': dic.get('authors', ''),
                    'Title': dic.get('title', ''),
                    'DOI': dic.get('doi', ''),
                    'Categories': dic.get('categories', ''),
                    'Abstract': dic.get('abstract', '')
                }
                data_list.append(data)
                count += 1
                # 如果已经达到了采样数量，则停止进一步读取
            if len(data_list) >= sample_size:
                break

    print(data_list[1000])
    # 创建DataFrame
    df = pd.DataFrame(data_list)

    # 分批写入Excel文件
    max_rows_per_file = 1000000  # 每个Excel文件的最大行数
    num_files = (len(df) // max_rows_per_file) + 1

    for i in tqdm(range(num_files), total=num_files, desc="Writing Excel"):
        start_row = i * max_rows_per_file
        end_row = min((i + 1) * max_rows_per_file, len(df))
        batch_df = df.iloc[start_row:end_row]
        file_path = f'{excel_file_path}_{i + 1}.xlsx'
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            batch_df.to_excel(writer, index=False)

import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed


def extract_technical_entities(abstract, system_prompt):
    try:
        # 调用 call_with_deep_jsonout 来处理 abstract
        response = call_with_deep_jsonout(system_prompt=system_prompt,
                                          question=f'Now please follow the requirements to start your work. abstract content: {abstract}')
        # print(f"-----------------------------\n{response}-----------------------------------")
        return response['technical_entities']
    except Exception as e:
        # 打印异常并返回空列表或适当的错误处理
        print(f"Error processing abstract: {e}")
        return []

def process_batch(batch, system_prompt, new_ws, start_row):
    results = [None] * len(batch)  # 用来存放结果，保证顺序一致
    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_index = {executor.submit(extract_technical_entities, abstract, system_prompt): idx
                           for idx, abstract in enumerate(batch)}
        for future in as_completed(future_to_index):
            idx = future_to_index[future]
            try:
                # 获取结果
                entities = future.result()
                # 将结果保存在对应的索引位置
                results[idx] = (batch[idx], entities)
            except Exception as e:
                print(f"Error retrieving result for abstract: {batch[idx]}, Error: {e}")
                results[idx] = (batch[idx], f"Error processing: {e}")

    # 保持顺序，将结果写入文件
    for abstract, entities in results:
        new_ws.append([abstract, str(entities)])

    return results

def main():
    # 读取原始xlsx文件
    input_file = r'E:\PaperAgent\archive/30wArxiv_1.xlsx'
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # 提示词
    system_prompt = '''Please carefully read the abstract, identify all technical entities and output them in JSON format. These entities may include, but are not limited to, technical terms, programming languages, frameworks, tools, algorithms, device names, evaluation metrics, etc. When extracting, please pay attention to the following points:

1. **Synonym Handling**: For different expressions of the same technical entity (such as abbreviations, full names, aliases, etc.), please unify them to the most commonly used or official name for output, and include the original expressions (such as abbreviations, full names, aliases, etc.) in `"other": []` after the official name.

2. **Disambiguation**: When encountering terms that may be ambiguous, please determine their specific meanings based on the context and extract them accurately. For example, "Java" could refer to the programming language or a specific technology or part of a framework; please determine its accurate meaning based on the abstract content.

3. **Completeness**: Ensure that the extracted technical entities are complete and accurate, without missing any important information.

4. **Standardization**: Maintain a unified format for output, listing each technical entity separately for ease of subsequent processing and analysis.

5. **Avoidance of Redundancy**: Ensure that within a single dialogue, no technical entity is repeated in the output. If multiple mentions of the same entity occur, only the first instance should be included in the output, and subsequent mentions should be omitted to maintain clarity and conciseness.

Please extract the technical entities from the abstract to the aforementioned requirements and output them in JSON format.

# Example Input:
"This paper presents a novel approach using Python and TensorFlow for deep learning. The proposed method utilizes the Keras API within TensorFlow to build and train convolutional neural networks. The performance of the model is evaluated using accuracy and F1 score as metrics. The results demonstrate that the proposed method outperforms traditional machine learning algorithms such as SVM and Random Forest."

# Example JSON Output:
{
  "technical_entities": [
    {
      "entity": "Python",
      "other": []
    },
    {
      "entity": "TensorFlow",
      "other": ["TF"]
    },
    {
      "entity": "Keras",
      "other": []
    },
    {
      "entity": "convolutional neural networks",
      "other": ["CNN"]
    },
    {
      "entity": "accuracy",
      "other": []
    },
    {
      "entity": "F1 score",
      "other": []
    },
    {
      "entity": "Support Vector Machine",
      "other": ["SVM"]
    },
    {
      "entity": "Random Forest",
      "other": []
    }
  ]
}
'''

    # 定义批次大小
    batch_size = 15000

    # 按行读取Abstract列的内容并进行处理
    abstracts = [row[4].strip().replace("\n", "") for row in
                 ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True) if row[4]]
    total_batches = len(abstracts) // batch_size + (1 if len(abstracts) % batch_size != 0 else 0)

    # 处理每个批次并保存到单独的文件
    for i in range(total_batches):
        batch = abstracts[i * batch_size:(i + 1) * batch_size]
        print(f"Processing batch {i + 1}/{total_batches}")

        # 为每个批次创建一个新的工作簿和工作表
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        new_ws.append(['Processed Abstract', 'Entities'])  # 添加标题

        # 处理批次并写入结果
        start_row = 2  # 从第二行开始写入新文件，因为第一行是标题
        results = process_batch(batch, system_prompt, new_ws, start_row)

        # 保存每个批次的结果到单独的文件
        batch_output_file = f'E:\\PaperAgent\\archive\\output_batch_{i + 1}.xlsx'
        new_wb.save(batch_output_file)

    print("Processing complete.")


if __name__ == "__main__":
    main()