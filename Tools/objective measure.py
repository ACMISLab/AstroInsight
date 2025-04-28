#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/12/25 21:55
# @Author : 桐
# @QQ:1041264242
# 注意事项：
import json
import urllib
import numpy as np
from openai import OpenAI
import requests
from FlagEmbedding import FlagAutoModel
import os
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from serpapi import GoogleSearch
import arxiv
import re

proxies ={
'http': 'http://localhost:7890',
'https': 'http://localhost:7890'
}
api_keys = [
    'ac824517e6059841fea33df130deabc6361320e29cc6462fa7c132f7cdca960b',
    '93d1ae55722302721a57fedaa1e5536eed20fdd3a6277f6dec1ff45fdf853d3f',
    'ddc784e89c413244cc243bfa513de217c94d8371c2650389b7e8754f105feca5',
    'b460716769b127258c8794f0faf2c2e170e0e945fc6f46da60efffd4d46f8b33',
    '6ee96172d22fc603a9488e654cc9dee5218050806e3202623fa62d91529f6f12',
    '79e5baf20f676076e6e98bda7241d19a5157d8c53fff1382314ec9d96fc58eb1',
    '8f8db0c95fbbf250e56aa25e3886963d390de3af54462558378caee5ddf0ed24',
    'cc44a0d23dd4f15fca9297684b3eafe989ae740425ffb41e30ca20993e1101c6',
    '5b71df6c18544ce4dec88c6711090b4b0481a4d0ae72bf07124e49733d2590a1',
    'daa68ab523c2c1f6bccc039e89246ab81dbae420b6ee2d095b584a00b6a561cf',
    '8f1f5bb1f30d2911df852ca0db26cd22f735f3f09e49e293ce917847fa9593c7',
    'ddb7812b74367181ac2a5d4ef9a177b6e2c0c7a8a4a4df3e23773f3a822c7602',
    '4e44d44d205ee21beefc916eea39f0dedcda2f4e99d23b3aa07d3d8f8efb17a0',
    '82c28a79894f355212190ec44b8d481a855f4d1e24128369942e32fe001e50d0',
    'e4d6fab6b6e8522cfe31c6d9de44b75f750aa283c235e59be823da6f45289c97',
    '566fef5dbc6c54d54727d15f59dd69823e1deab36cfb0809ab94e37ca61620f0',
]
Over=False


def find_last_year(s):
    # 使用正则表达式查找所有看起来像年份的数字（即1000-9999之间的四位数）
    matches = re.findall(r'\b([1-9]\d{3})\b', s)
    if not matches:
        return None  # 如果没有找到任何匹配项，则返回None
    # 返回最后一个匹配项，假设它是所需的年份
    return matches[-1]

def deal_year():
    # 加载已有的工作簿和选择工作表
    file_path = 'E:\PaperAgent/相关文献统计.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # 遍历工作表的所有行
    index=1
    for row in ws.iter_rows(min_row=2,values_only=True):
        index+=1
        year=find_last_year(row[5])
        print(year)

        ws.cell(row=index, column=11, value=year)

    wb.save(file_path)

def deal_ori_abs():
    # 加载已有的工作簿和选择工作表
    file_path = 'E:\PaperAgent/待评估数据客观实验.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    index=-1
    num=0
    for row in ws.iter_rows(min_row=2,values_only=True):
        index += 1
        if index%3!=0:
            continue
        num+=1
        content=row[2]
        match = re.search(fr'### Paper abstract:(.*?)(?=###|\Z)', content, re.DOTALL)
        abstract = match.group(1).strip()

        print(row[0])
        # print(abstract)
        ws.cell(row=num+1, column=9, value=row[0])
        ws.cell(row=num+1, column=10, value=abstract)

    print(index)

    wb.save(file_path)

def remove_non_english(text):
    # 使用正则表达式只保留英文字母
    cleaned_text = re.sub(r'[^a-zA-Z ]', '', text)
    return cleaned_text

def get_daily_papers(query="astronomy", max_results=5):
    """
    @param topic: str
    @param query: str
    @return paper_with_code: dict
    """
    paper_titles=[]
    paper_abstracts=[]
    search_engine = arxiv.Search(
        query=query,
        max_results=max_results,
        sort_by=arxiv.SortCriterion.Relevance
    )

    for result in search_engine.results():
        paper_titles.append(result.title)
        paper_abstracts.append(result.summary.replace("\n", " "))

    return paper_titles,paper_abstracts

def search_google_scholar(doi):
    proxies = {
        'http': 'http://localhost:7890',
        'https': 'http://localhost:7890'
    }

    # Google Scholar的搜索URL
    search_url = f"https://scholar.google.com/scholar?q={urllib.parse.quote(doi)}"

    print(search_url)
    # 模拟请求
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }

    response = requests.get(search_url, headers=headers , proxies=proxies)

    # print(response)

    soup = BeautifulSoup(response.text, 'html.parser')

    # print(soup)

    # 查找包含[PDF]文本的链接
    pdf_links = []
    cite_count=0
    erro=True
    abstract=''
    others=''

    for link in soup.find_all('a'):

        # if '[PDF]' in link.get_text():
        #     pdf_links.append(link['href'])

        if 'Cited by ' in link.get_text():
            cite_count=link.text[8:]
            erro=False


    try:
        abstract=soup.find_all(attrs={'class':'gs_fma_snp'})[0].text
    except:
        erro=True

    try:
        others = soup.find_all(attrs={'class': 'gs_a gs_fma_p'})[0].text
    except:
        erro=True

    print(cite_count)
    print(others)
    print(abstract)
    print(erro)
    return abstract,cite_count,others,erro

def Serp_google(query='',api_key=''):

    params = {
        "engine": "google_scholar",
        # "author_id": "LSsXyncAAAAJ",
        "q": query,
        "api_key": api_key
    }

    search = GoogleSearch(params)
    results = search.get_dict()

    print(results)

    try:
        if results['error']=='Your account has run out of searches.' or results['error']=='Your account has been throttled. You are exceeding 100 searches per hour. Please upgrade your plan, spread out your searches, or contact support.':
            global Over
            Over=True
            return '','','','','','','',''
    except:
        pass

    # print(results)
    print('--------------------------------------------------------------------------------')
    # print(results)
    # articles = results["articles"]
    # citations = results["citations"]
    # cited_by = results["cited_by"]
    # citation = results["citation"]
    try:
        organic_results = results["organic_results"]
    except:
        return '', '', '', '', '', '', '', 'BIG ERRO'

    # 将 organic_results 转换为 JSON 格式的字符串
    organic_results_json = json.dumps(organic_results, ensure_ascii=False, indent=4)
    # print(organic_results_json)

    raw_link=results['search_metadata']['raw_html_file']
    title=organic_results[0]['title']
    link = organic_results[0]['link']
    abs= organic_results[0]['snippet']
    publication_info = organic_results[0]['publication_info']['summary']
    try:
        pdf = organic_results[0]['resources'][0]['link']
    except:
        pdf = ''
    try:
        citeby = organic_results[0]['inline_links']['cited_by']['total']
    except:
        citeby=0

    # print(f'title_search:{title}\nabs:{abs}\ncite:{citeby}\nothers:{publication_info}\nlink:{link}\nPdf:{pdf}\nRaw_link:{raw_link}')

    # 模拟请求
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    response = requests.get(raw_link, headers=headers)
    soup = BeautifulSoup(response.text, 'html.parser')
    try:
        abstract = soup.find_all(attrs={'class': 'gs_fma_snp'})[0].text
        erro=False
    except:
        erro=True
        abstract=''
    # print(abstract)

    return title,citeby,publication_info,link,pdf,raw_link,abstract,erro

def Serp_google_save():
    global Over
    # 加载已有的工作簿和选择工作表
    file_path = 'E:\PaperAgent/相关文献统计.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    index=0
    api=0
    # 遍历工作表的所有行
    for row in ws.iter_rows(min_row=1954,values_only=True):
        flag=False
        # 这里假设我们想要打印每一行的数据
        topic=row[0]
        title=(row[1])
        print(f'{index} {topic}')

        save_path = 'E:\PaperAgent/test.xlsx'
        wsave = load_workbook(filename=save_path)
        wsok = wsave.active
        # 获取当前最后一行的行号
        last_row = wsok.max_row
        for sub_row in wsok.iter_rows(values_only=True):
            if sub_row[1]==title:
                new_data=[topic,sub_row[1],sub_row[2],sub_row[3],sub_row[4],sub_row[5],sub_row[6],sub_row[7],sub_row[8],sub_row[9]]
                flag=True
                break

        if flag==False:
            title_search, citeby, publication_info, link, pdf, raw_link, abstract,erro=Serp_google(query=title[:-4],api_key=api_keys[5+api])#9
            if Over==True:
                Over = False
                api += 1
                print(f'api_keys:{api_keys[5 + api]}')
                title_search, citeby, publication_info, link, pdf, raw_link, abstract, erro = Serp_google(query=title[:-4], api_key=api_keys[5 + api])  # 9
                new_data = [topic, title, title_search, abstract, citeby, publication_info, link, pdf, raw_link, erro]
            else:
                new_data=[topic,title,title_search,abstract,citeby,publication_info,link,pdf,raw_link,erro]

        print(new_data)

        last_row += 1
        for col_num, value in enumerate(new_data, start=1):
            wsok .cell(row=last_row, column=col_num, value=value)
        # 保存更改
        wsave.save(save_path)

def Serp_google_raw_abs():

    # 加载已有的工作簿和选择工作表
    file_path = 'E:\PaperAgent/相关文献统计.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # 遍历工作表的所有行
    index=1
    for row in ws.iter_rows(min_row=2,values_only=True):
        index+=1
        if row[9]== True:
            title=row[1][:-4]

            paper_title=''
            paper_titles,paper_abstracts=get_daily_papers(query=title)
            print('-----------------------------------------------------')
            #print(title)
            #print(row[1][:-4])
            #print(paper_title)
            # print(paper_abstract)

            for num,paper_title in enumerate(paper_titles):
                if  remove_non_english(title)== remove_non_english(paper_title) or  remove_non_english(paper_title)== remove_non_english(row[1][:-4]):
                    print(remove_non_english(title))
                    print(remove_non_english(row[1][:-4]))
                    print(remove_non_english(paper_title))
                    print('over')
                    ws.cell(row=index, column=10, value=False)
                    ws.cell(row=index, column=4, value=paper_abstracts[num])
                    break
                    # 保存更改
                # else:
                #     print(remove_non_english(title))
                #     print(remove_non_english(row[1][:-4]))
                #     print(remove_non_english(paper_title))
                #     print('ok')
                #     ws.cell(row=index, column=4, value=paper_abstract)

    wb.save(file_path)

# 计算核心论文摘要与相关论文摘要的相似度
def calculate_similarity(related,target,open_eu=False):

    # 加载模型
    model = FlagAutoModel.from_finetuned(r'E:\PaperAgent\bge-large-en-v1.5',
                                         # query_instruction_for_retrieval="Represent this sentence for searching relevant passages:",
                                         use_fp16=True)
    # 计算嵌入
    embeddings_1 = model.encode(related)
    # print(len(embeddings_1))
    embeddings_2 = model.encode(target)
    # print(embeddings_1)
    # print(embeddings_2)

    # 计算相似度
    similarity = embeddings_1 @ embeddings_2.T

    if open_eu==True:
        # 计算欧几里得距离
        euclidean_distance = np.linalg.norm(embeddings_1 - embeddings_2)
        # print(similarity)
        # print(euclidean_distance)
        return similarity, euclidean_distance
    else:
        # print(similarity)
        return similarity

def contemporary_impact():
    num = 0
    # 指定路径
    dir = 'E:\PaperAgent\paper'
    # 获取该路径下所有文件和文件夹的名字列表
    topic_dirs = os.listdir(dir)

    for path in topic_dirs:
        print(num)
        num+=1
        print(path)
        paper_path=dir+f'/{path}/pdf'
        papers=os.listdir(paper_path)

        for paper in papers:
            # 加载已有的工作簿和选择工作表
            file_path = 'E:\PaperAgent/相关文献.xlsx'
            wb = load_workbook(filename=file_path)
            ws = wb.active  # 或者指定工作表名：ws = wb['Sheet1']

            print(paper)
            # print(paper[:-4])
            try:
                abstract,cite_count,others,erro=search_google_scholar(doi=paper[:-4])
                # 假设你有如下 DataFrame 要追加
                new_data = [path,paper,abstract,cite_count,others,erro]
            except:
                new_data=[path,paper,False]

            # 获取当前最后一行的行号
            last_row = ws.max_row
            last_row += 1

            for col_num, value in enumerate(new_data, start=1):
                ws.cell(row=last_row, column=col_num, value=value)

            # 保存更改
            wb.save(file_path)

def start_objective_exp():
    # 加载已有的工作簿和选择工作表
    file_path = 'E:\PaperAgent/客观实验数据_OUR_ORI.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active

    file_path = 'E:\PaperAgent/客观实验数据_OUR.xlsx'
    wbI = load_workbook(filename=file_path)
    wsI = wbI.active

    #检查代码
    # for row in ws.iter_rows(min_row=2,max_row=62, values_only=True):
    #     flag=False
    #     for sub_row in wsI.iter_rows(min_row=2, values_only=True):
    #         if row[0]==sub_row[0]:
    #             flag=True
    #             break
    #     if flag!=True:
    #         print(row[0])

    index=1
    for sub_row in wsI.iter_rows(min_row=2, values_only=True):
        index+=1
        for row in ws.iter_rows(min_row=2, max_row=62, values_only=True):
            if sub_row[0]==row[0]:
                similarity, euclidean_distance=calculate_similarity(related=sub_row[3], target=row[1], open_eu=False)
                print(f'index:{index}, name:{sub_row[0]}, similarity:{similarity}, ed:{euclidean_distance}')
                wsI.cell(row=index, column=12, value=similarity)
                wsI.cell(row=index, column=13, value=similarity)
                break

    wbI.save('E:\PaperAgent/客观实验数据_OUR.xlsx')

        # break

def llm_ablation():
    file_path = r'E:\PaperAgent\结果/待评估数据_LLM_Ablation.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active

    Key = OpenAI(api_key="sk-80cc66e836004e6ca698eb35206dd418", base_url="https://api.deepseek.com/v1")
    client = Key
    index=0

    # 遍历工作表的所有行 4 5
    for row in ws.iter_rows(min_row=2,values_only=True):
        index+=1
        if index % 4 == 1:

            user_prompt=f"""# Now,please start your work. 
# The Draft 1 is :
{row[7]}
# The Draft 2 is :
{row[9]}
    
# Output format:
Rationale: [Explain the improvements (if any), their quality, and significance, and justify your score]
Score: [Your score from -5 to 5]"""

            response = client.chat.completions.create(
                model="deepseek-chat",
                temperature=0.7,
                messages=[
                    {"role": "system",
                     "content": "You are a strict expert reviewer. There are two research drafts on a specific topic: Draft 1 and Draft 2. Please compare Draft 2 with Draft 1 and assess whether Draft 2 improves upon Draft 1. If Draft 2 shows positive improvements, award positive points. If the improvements are negative, award negative points. If there are no significant improvements, award 0 points. Please score based on the quality and significance of the improvements, with a scale from [-5, 5]."},
                    {"role": "user", "content": user_prompt},
                ]
            )

            print(response.choices[0].message.content)
            ws.cell(row=index+1, column=11, value=response.choices[0].message.content)
            print(f'-----------------------------over------------------------------')
    wb.save(file_path)

def get_llm_ablation_score():
    file_path = r'E:\PaperAgent\结果/待评估数据_LLM_Ablation_Score.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active

    index = 0

    # 遍历工作表的所有行 4 5
    for row in ws.iter_rows(min_row=2, values_only=True):
        index += 1
        if index % 4 == 1:

            # 正则表达式：提取"Score:"后第一次出现的数字
            pattern = r'Score:\s*([+-]?\d+)'

            # 使用re.search()查找匹配
            match = re.search(pattern, str(row[10]))

            if match:
                # 输出提取的数字
                print("提取的数字是:", match.group(1))
            else:
                print("未找到匹配的数字",index+1)

            ws.cell(row=index + 1, column=11, value=match.group(1))
            print(f'-----------------------------over------------------------------')
    wb.save(file_path)



# calculate_similarity(related="hello,man",target="hello!!! man",open_eu=True)
# contemporary_impact()
# search_google_scholar(doi='AI governance and ethics framework for sustainable AI and sustainability')
# Serp_google()
# Serp_google_save()
# Serp_google_raw_abs()
# deal_year()
# deal_ori_abs()
# start_objective_exp()
# get_llm_ablation_score()