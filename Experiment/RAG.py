#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/12/31 14:03
# @Author : 桐
# @QQ:1041264242
# 注意事项：
# 加载已有的工作簿和选择工作表
import json
import shutil
import requests
from openai import OpenAI
from openpyxl import load_workbook
import os

def save_string_to_markdown(content, file_path):
    # 确保目录存在，如果不存在则创建
    directory = os.path.dirname(file_path)
    if not os.path.exists(directory):
        os.makedirs(directory)

    # 写入文件
    try:
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write(content)
        # print(f"内容已成功写入 {file_path}")
    except Exception as e:
        print(f"写入文件时发生错误: {e}")
def get_abs():
    file_path = 'E:\PaperAgent/客观实验数据_OUR.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # 遍历工作表的所有行
    for row in ws.iter_rows(min_row=2, values_only=True):
        topic=row[0]
        title=row[1][:-4]
        abs=row[3]
        directory_path=fr'E:\PaperAgent\paper/{topic}'
        if os.path.exists(directory_path) and os.path.isdir(directory_path):
            save_string_to_markdown(str(abs), fr'E:\PaperAgent\paper/{topic}/abs/{title}_abs.md')
        else:
            print(topic)
        # break
def copy_file():
    dst_file= r'E:\PaperAgent\all_md/'
    file_path = 'E:\PaperAgent/客观实验数据_OUR_ORI.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # 遍历工作表的所有行
    for row in ws.iter_rows(min_row=2,max_row=62, values_only=True):
        topic=row[0]
        directory_path = fr'E:\PaperAgent\paper/{topic}'
        if os.path.exists(directory_path) and os.path.isdir(directory_path):
            for file in os.listdir(directory_path+'/markdown'):
                shutil.copy2(directory_path+f'/markdown/{file}', dst_file)
            for file in os.listdir(directory_path+'/abs'):
                shutil.copy2(directory_path+f'/abs/{file}', dst_file)

        else:
            print(topic)
        # break

def exp_start():

    file_path = 'E:\PaperAgent/客观实验数据_OUR_ORI.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active
    # 遍历工作表的所有行
    num=1
    for row in ws.iter_rows(min_row=2,max_row=62, values_only=True):
        num+=1
        topic=row[0]
        query=topic
        data = {
          "query": query,
          "knowledge_base_name": "exp",
          "top_k": 100,
          "score_threshold": 1,
          "file_name": "",
          "metadata": {}
        }
        # 设置请求头
        headers = {
            'Content-Type': 'application/json'
        }
        # 发送 POST 请求
        response = requests.post(url='http://210.40.16.12:24438/knowledge_base/search_docs', data=json.dumps(data), headers=headers)
        paper_prompt=''
        # print(response.text)
        for index,infor in enumerate(json.loads(response.text)):
            # print(infor['page_content'])
            paper_prompt+=f"Releated paper information {index+1}:\"{infor['page_content']}\"\n"
        # print(paper_prompt)

        user_prompt = f"""# Task Definition\nYou are going to generate a research draft that should be original, clear, feasible, relevant, and significant to astronomy field. This will be based on some related paper information.\n
    
I will provide the related paper information as follows:
# Related Paper Information
{paper_prompt}
# Output Requirements
With the provided related paper information, your objective now is to formulate a research draft that not only builds upon these existing studies but also strives to be original, clear, feasible, relevant, and significant. In addition, you also need to provide a detailed Rationale、Necessary technical details、possible datasets to ultimately form the paper title,abstract,methods and experiments. Note: The final title should be appealing, and the experimental content should include relevant baselines as much as possible.
    
# Respond in the following format:
### Problem:
### Rationale:
### Necessary technical details:
### Datasets:
### Paper title:
### Paper abstract:
### Methods:
### Experiments:"""

        print(f'-----------------------------{topic}------------------------------')
        print(user_prompt)

        Key = OpenAI(api_key="sk-80cc66e836004e6ca698eb35206dd418", base_url="https://api.deepseek.com/v1")
        client = Key
        response = client.chat.completions.create(
            model="deepseek-chat",
            temperature=0.7,
            messages = [
                {"role": "system", "content": "You are a research expert in astronomy and computer whose primary goal is to identify promising, new, and key scientific problems based on existing scientific literature, in order to aid researchers in discovering novel and significant research opportunities that can advance the field."},
                {"role": "user", "content": user_prompt},
            ]
        )

        print(response.choices[0].message.content)
        ws.cell(row=num, column=3, value=response.choices[0].message.content)
        print(f'-----------------------------over------------------------------')
    wb.save(file_path)

def exp_table():
    file_path = r'E:\PaperAgent\结果/待评估数据.xlsx'
    wb = load_workbook(filename=file_path)
    ws = wb.active

    file_path = r'E:\PaperAgent\结果/待评估数据（Acmis）.xlsx'
    wbg = load_workbook(filename=file_path)
    wsg = wbg.active
    index=0
    add=0
    for row in wsg.iter_rows(min_row=2, values_only=True):
        index+=1

        iteration_1=row[6]
        iteration_2=row[8]
        iteration_3=row[10]
        Novelty=row[14]
        Feasibility=row[15]
        Rationale=row[16]
        Technical=row[17]
        Dataset=row[18]
        Methodology=row[19]
        Experimental=row[20]
        Experience=row[21]
        Reviewer=row[22]
        # print(f'{iteration_1} {iteration_2} {iteration_3} {Novelty} {Experience} {Reviewer}')

        ws.cell(row=index+1+add, column=7, value=iteration_1)
        ws.cell(row=index+1+add, column=9, value=iteration_2)
        ws.cell(row=index + 1 + add, column=11, value=iteration_3)
        ws.cell(row=index + 1 + add, column=15, value=Novelty)
        ws.cell(row=index + 1 + add, column=16, value=Feasibility)
        ws.cell(row=index + 1 + add, column=17, value=Rationale)
        ws.cell(row=index + 1 + add, column=18, value=Technical)
        ws.cell(row=index + 1 + add, column=19, value=Dataset)
        ws.cell(row=index + 1 + add, column=20, value=Methodology)
        ws.cell(row=index + 1 + add, column=21, value=Experimental)
        ws.cell(row=index + 1 + add, column=22, value=Experience)
        ws.cell(row=index + 1 + add, column=23, value=Reviewer)

        if index % 3 == 0:
            add+=1

    # 保存更改
    wb.save(r'E:\PaperAgent\结果/ACMIS_RAG增补实验.xlsx')




# copy_file()
# exp_start()
exp_table()