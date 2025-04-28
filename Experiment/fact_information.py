#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/10/21 22:39
# @Author : 桐
# @QQ:1041264242
# 注意事项：
import openpyxl
from scipy.stats import spearmanr

from PaperAgent.API.LLM_API import call_with_deep
from openpyxl import load_workbook
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
import nltk
import re
import scipy.stats as stats
import matplotlib.pyplot as plt
import seaborn as sns
import pandas as pd
from sklearn.metrics import cohen_kappa_score

def remove_number_prefix(sentence):
    # 定义一个正则表达式模式，用于匹配句子开头的数字和随后的句点空格
    pattern = r'^\d+\. '
    # 利用re.sub函数，将匹配到的部分替换为空字符串，以此移除它
    modified_sentence = re.sub(pattern, '', sentence)
    return modified_sentence

def fact_information_extraction():
    # 加载现有的 XLSX 文件
    file_path = r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/fact_information_experiment.xlsx"  # 替换为你的文件路径
    wb = load_workbook(filename=file_path)
    ws = wb.active

    system_prompt = f"""# Task Definition: Please extract the objective factual information from the following provided paper title and abstract, and output it. Ensure that the extracted information meets the following requirements:

1. **Avoid Demonstrative Pronouns**: Do not use demonstrative pronouns such as 'this', 'that', 'these', 'those', 'this work', 'this survey', 'the study', etc. Instead, use 'Related research'. Ensure that each piece of information is self-contained and clear.
2. **Dataset Description Standards**: When describing datasets, not only specify which dataset was used, but also clearly state the source of the dataset and its intended purpose (e.g., The data comes from [data source] can be used in [research topic]. ).
3. **Result Presentation Standards**: For experimental results, avoid using ambiguous phrases like "the results indicate" and instead use "Related research show [experimental results]," ensuring that the content of the experimental results is clear and specific.
4. **Accuracy**: All information should be an explicit statement from the abstract, without personal speculation or interpretation.
5. **Completeness**: Include all main factual points from the abstract, avoiding the omission of key information.
6. **Conciseness**: Each fact should be a concise and clear statement.
7. **Structured Format**: Present each factual information in a numbered list format, with each fact on a separate line.
8. **Independence and Clarity**: Each piece of output information must be able to stand alone without context, with clear meaning, free from ambiguity or unclear references.

# Example output format:

1.Related research have shown that [research topic].
2.Related research have used [methods], and find that [experimental results].
3.The data comes from [data source] can be used in [research topic].
4.Related research have shown that [experimental results].
5.Related research have found that [conclusion]."""

    for index, paper in enumerate(ws.iter_rows(min_row=2,max_row=63, values_only=False)):
        title = paper[1].value
        abstract = paper[2].value

        # print(abstract)
        user_prompt = f"""Now,please following these rules to extract the factual information from following paper:\ntitile:{title}\nabstract:\n{abstract}\n"""

        result = call_with_deep(system_prompt=system_prompt, question=user_prompt, temperature=0.01)

        # 将结果写入到第5列（索引为4）
        paper[4].value = result
        print(title)
        # break

        # result = call_with_deep(system_prompt=system_prompt, question=user_prompt,temperature=0.01)

    # 保存工作簿
    wb.save(r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/result.xlsx")

def experiment_score():

    # 加载现有的 XLSX 文件
    file_path = r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/数据标记.xlsx"  # 替换为你的文件路径
    wb = load_workbook(filename=file_path)
    ws = wb.active

    smoothie = SmoothingFunction().method4  # 用于平滑处理

    for index, paper in enumerate(ws.iter_rows(min_row=2, max_row=60, values_only=False)):

        # 计算bleu
        reference_paragraph = paper[5].value

        candidate_paragraph = paper[6].value

        # 将段落分词为词列表
        reference_tokens = nltk.word_tokenize(reference_paragraph)
        candidate_tokens = nltk.word_tokenize(candidate_paragraph)

        # 使用BLEU计算相似性
        bleu_score = sentence_bleu([reference_tokens], candidate_tokens, smoothing_function=smoothie)

        paper[14].value = bleu_score

        # print(f"bleu_score:{bleu_score}")

        #--------------------------
        #计算recall
        reference_paragraph_split=reference_paragraph.split('\n')
        candidate_paragraph_split=candidate_paragraph.split('\n')

        all=0
        hit=0

        for reference_sentence in reference_paragraph_split:
            reference_sentence=remove_number_prefix(reference_sentence)
            # print(reference_sentence)
            all+=1
            for candidate__sentence in candidate_paragraph_split:
                candidate__sentence=remove_number_prefix(candidate__sentence)
                bleu_score = sentence_bleu([nltk.word_tokenize(reference_sentence)], nltk.word_tokenize(candidate__sentence), smoothing_function=smoothie)
                if bleu_score>=0.5:
                    hit+=1

        recall=hit/all

        paper[15].value = recall

        # print(f"recall:{recall}")

        #计算recall
    # 保存工作簿
    wb.save(r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/result_score_test.xlsx")

def quantity_score():

    # 加载现有的 XLSX 文件
    file_path = r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/数据标记.xlsx"  # 替换为你的文件路径
    wb = load_workbook(filename=file_path)
    ws = wb.active

    for index, paper in enumerate(ws.iter_rows(min_row=2, max_row=60, values_only=False)):

        # 计算bleu
        reference_paragraph = paper[5].value

        candidate_paragraph = paper[6].value

        user_prompt=f"""# Task Definition:
You are required to score the model annotation results based on the human annotation outcomes. Because, when extracting factual information from paper abstracts, it is necessary to compare the performance of model annotations against human annotations. This comparison will be quantified through a scoring criteria designed to evaluate the differences between the two. The scoring will be based on the completeness, accuracy, and clarity of the information.

# Scoring Criteria:

1. Completeness:
- Full Marks Standard: The model annotation covers all key information points present in the human annotation, with no omissions.
- Scoring Details:
Excellent (5 points): Fully covers all information points from the human annotation.
Good (4 points): Covers most information points but has slight omissions.
Average (3 points): Covers more than half of the information points but has significant omissions.
Poor (2 points): Covers only a few information points.
Very Poor (1 point): Covers almost no information points.

2. Accuracy:
- Full Marks Standard: The information in the model annotation is identical to that in the human annotation, with no errors or misleading content.
- Scoring Details:
Excellent (5 points): Information is completely accurate, with no errors.
Good (4 points): Information is mostly accurate but has a few minor errors or differences in expression.
Average (3 points): Information has some accuracy but contains multiple minor errors or a few key errors.
Poor (2 points): Information has many errors but still contains some correct content.
Very Poor (1 point): Information is almost entirely incorrect or severely misleading.

3. Clarity:
- Full Marks Standard: The information in the model annotation is clearly expressed, well-organized, and easy to understand.
- Scoring Details:
Excellent (5 points): Expression is very clear, with good logical structure.
Good (4 points): Expression is clear, but there may be some parts that are slightly verbose or not as concise.
Average (3 points): Expression is generally clear but has some parts that are vague or difficult to understand.
Poor (2 points): Expression is confusing, making it hard to grasp the main information.
Very Poor (1 point): Expression is extremely confusing, making it almost impossible to extract useful information.

# Output Format Example:
Completeness: The model annotation covers 80% of the information points in the human annotation but misses two key details. → Score: 4 points (Good)
Accuracy: Most of the information in the model annotation matches the human annotation, but there is a discrepancy in one key data point. → Score: 4 points (Good)
Clarity: The model annotation is overall clear, but one section is slightly verbose and could be more concise. → Score: 4 points (Good)

Overall Evaluation: The model annotation performs well in completeness, accuracy, and clarity, but there is still room for improvement, particularly in the comprehensive coverage of information and the accuracy of key data. Therefore, the overall score is 4 points (Good).

# Please use the above scoring criteria to ensure the fairness and effectiveness of the evaluation. Now,start your work:

The human annotations result:'''{reference_paragraph}'''

The model annotations result:'''{candidate_paragraph}'''
"""

        result=call_with_deep(system_prompt="You are a helpful assistant",question=user_prompt)

        # 将结果写入到第5列（索引为4）
        paper[16].value = result

    # 保存工作簿
    wb.save(r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/实验.xlsx")

def extra_score():

    # 加载现有的 XLSX 文件
    file_path = r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/数据标记.xlsx"  # 替换为你的文件路径
    wb = load_workbook(filename=file_path)
    ws = wb.active

    for paper in ws.iter_rows(min_row=2, max_row=60, values_only=False):
        try:
            # 使用正则表达式提取得分
            scores = re.findall(r'Score: (\d+) points', paper[16].value)
            print(scores)

            paper[1].value=scores[0]
            paper[2].value=scores[1]
            paper[3].value=scores[2]
            paper[4].value=scores[3]
            print("ok")

        except:
            pass

    # 保存Excel文件
    wb.save(r'C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/scores.xlsx')

def Kappa_class(score):
    if score<0.2 and score>=0:
        return 1
    elif score<0.4 and score>=0.2:
        return 2
    elif score<0.6 and score>=0.4:
        return 3
    elif score<0.8 and score>=0.6:
        return 4
    elif score<=1.0 and score>=0.8:
        return 5

def Kappa():
    import numpy as np

    # 加载现有的 XLSX 文件
    file_path = r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Experiment/最终数据v1.0.xlsx"  # 替换为你的文件路径
    wb = load_workbook(filename=file_path)
    ws = wb.active

    human_scores=[]
    model_scores=[]

    for paper in ws.iter_rows(min_row=2, max_row=36, values_only=False):
        model_scores.append(float(paper[10].value))
        human_scores.append(float(paper[19].value))

        # 初始相关系数
        initial_corr, _ = spearmanr(human_scores, model_scores)
        print(initial_corr)

        # 存放每个点去除后的相关系数
        correlations = []

        for i in range(len(human_scores)):
            temp_human_scores = human_scores[:i] + human_scores[i + 1:]
            temp_model_scores = model_scores[:i] + model_scores[i + 1:]

            corr, _ = spearmanr(temp_human_scores, temp_model_scores)
            correlations.append((corr, i))

            # 按相关系数变化排序（从最小到最大）
        correlations.sort(key=lambda x: abs(initial_corr - x[0]))

        # 输出影响最小和最大的点
        print("对整体相关性影响最小的点（按影响从小到大排序）：")
        for corr, index in correlations:
            print(f"去除第{index + 1}个点后的相关系数: {corr}，变化: {abs(initial_corr - corr)}")

            # 输出排序后的数据点
        sorted_indices = [index for corr, index in correlations]
        sorted_human_scores = [human_scores[i] for i in sorted_indices]
        sorted_model_scores = [model_scores[i] for i in sorted_indices]

        print("排序后的人类评分:", sorted_human_scores)
        print("排序后的模型评分:", sorted_model_scores)

    # 初始相关系数
    initial_corr, _ = spearmanr(human_scores, model_scores)
    print(initial_corr)
    #
    # # 计算肯德尔等级相关系数
    # kendall_corr, _ = kendalltau(human_scores, model_scores)
    # print(f"肯德尔等级相关系数: {kendall_corr}")
    #
    # # 计算均方误差（MSE）
    # mse = np.mean((human_scores - model_scores) ** 2)
    # print(f"均方误差: {mse}")
    #
    # # 计算均方根误差（RMSE）
    # rmse = np.sqrt(mse)
    # print(f"均方根误差: {rmse}")
Kappa()
# extra_score()
# experiment_score()
# fact_information_extraction()

# bleu()
