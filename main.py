#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time : 2024/7/7 21:48
# @Author : 桐
# @QQ:1041264242
# 注意事项：
import os
import random
import re

import openpyxl

from PaperAgent.API.LLM_API import call_with_deep,call_with_deep_jsonout,call_with_qwenmax,call_with_qwenplus,call_with_DeepSeek_R1_250120
from PaperAgent.API import LLM_API
from PaperAgent.API.ADS_API import search_title_DataAndCite
from PaperAgent.API.arXiv_API import search_paper,download_pdf,sanitize_folder_name
from PaperAgent.API.WikiSearch import search,get_description
from PaperAgent.API.GoogleSearch_API import download_all_pdfs
from PaperAgent.API.PdfToMd_API import pdf2md,pdf2md_mineruapi
from PaperAgent.Tools.GetKeyWordScore import SearchKeyWordScore
from PaperAgent.Tools.Doc2X import pdf2md_docx2
from openpyxl import load_workbook,Workbook
from Tools.PaperAgent_MOA import moa_idea_iteration,moa_model,moa_table
import ast

save_exp=r"xxxxxxxxx"


def read_markdown_file(file_path):
    """
    读取指定Markdown文件的内容，并将其打印到控制台。

    参数:
    file_path (str): Markdown文件的路径。
    """

    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
            # print(content)
            return content

    except FileNotFoundError:
        print(f"\033[1;31m | ERRO     | file can't find: {file_path} \033[0m")
    except IOError as e:
        print(f"\033[1;31m | ERRO     | load file erro: {e} \033[0m")

def update_excel_row(file_path,target_topic, new_data, id_row,sheet_name='Sheet1'):
    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)

    # 选择工作表
    sheet = workbook[sheet_name]

    # 遍历所有行
    for row in sheet.iter_rows(min_row=2, values_only=False):  # 从第二行开始，忽略标题行
        if row[2].value == target_topic:  # 假设"Topic"是第一列
            row[id_row].value = new_data  # 假设新数据从第二列开始写入
            break  # 找到目标行后停止循环
    # 保存更改
    workbook.save(file_path)

def information_compression(doi,title,topic):
    paper_pdf_path= download_all_pdfs(dois=doi,title=title,topic=topic)

    # print(paper_pdf_path)

    task_id = pdf2md_mineruapi(file_path=paper_pdf_path, topic=topic)
    # task_id=pdf2md(paper_pdf_path,topic)
    # task_id = pdf2md_docx2(paper_pdf_path)

    pattern1 = r'## References.*' #docx
    pattern2 = r'# References.*'

    # 读取markdown内容
    if task_id != 0:
       paper_content = read_markdown_file(file_path=fr"E:\PaperAgent\paper\{topic}\markdown/{task_id}.md")
       paper_content = re.sub(pattern1, '', paper_content, flags=re.DOTALL)
       paper_content = re.sub(pattern2, '', paper_content, flags=re.DOTALL)
    # print(paper_content)

    # print(paper_content)

    system_prompt="""# Task Definition
Please compress the content of the provided academic paper according to the following requirements:
1.The Methods and Experiments sections should be described in extreme detail, as they are the most crucial parts of the compression. Ensure no key technical details, experimental steps, or relevant formulas are omitted.
2.While compressing, ensure the scientific accuracy and integrity of the information are maintained.
3.Avoid redundant or repetitive content, focusing on extracting the most critical points.
4.Ensure that the overall logic and flow of the paper remain clear and coherent.
5.Ignore the Abstract Do not include or mention any part of the abstract section of the paper.

# Output Requirements
## Introduction
Provide an overview of the research background, purpose, and significance.
Highlight the research motivation, main research questions, and objectives.
Extract the most critical points, avoiding lengthy background information.

## Methods
This section should be extremely detailed.
Summarize the methodology used in the research.
Describe the research design, data sources, sample selection, experimental setup, and technical details, ensuring the methodology section remains complete.
Retain all relevant formulas, algorithms, and mathematical expressions to maintain scientific rigor.
If certain details, like sample selection or data sources, are not explicitly mentioned in the paper, feel free to skip them rather than forcing their inclusion.

## Experiments/Results
This section should be extremely detailed.
Provide a highly condensed summary of the experimental results, focusing on the main findings and their alignment (or discrepancies) with the hypotheses.
Preserve the essential elements of the experimental design, including the purpose of the experiments, hypotheses, or research questions.
Only retain the key experimental steps and procedures, while simplifying secondary details.
Ensure all critical data and results are kept intact to maintain the accuracy and completeness of the data.
Clearly indicate the baselines used in the paper.

## Conclusions
Distill the main conclusions of the study, emphasizing the research findings' contributions to the field and how they address the research questions.
Highlight the novelty of the research and its implications for future work.

## Future Work
Indicate possible directions for future research based on the current study’s results.
Discuss the limitations of the current research and provide suggestions for improvement."""

    #信息压缩：use old
    # compression_result = call_with_deep(question=f"The content is '''{paper_content}'''",system_prompt=system_prompt)

    #use r1
    compression_result = call_with_DeepSeek_R1_250120(question=f"The content is '''{paper_content}'''", system_prompt=system_prompt)

    directory = fr"E:\PaperAgent\paper\{topic}\compression"
    if not os.path.exists(directory):
        os.makedirs(directory)

    with open(fr"{directory}/{task_id}.md",'w',encoding='utf-8') as f:
        f.write(compression_result)

    # with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\compression/{task_id}.md",'w',encoding='utf-8') as f:
    #     f.write(compression_result)

    # print(compression_result)

    return compression_result

def information_compression2(paper_pdf_path,topic):

    task_id = pdf2md_mineruapi(paper_pdf_path, topic)
    # task_id=pdf2md(paper_pdf_path,topic)
    # task_id = pdf2md_docx2(paper_pdf_path)

    pattern1 = r'## References.*'
    pattern2 = r'# References.*'

    # 读取markdown内容
    if task_id != 0:
       paper_content = read_markdown_file(file_path=fr"E:\PaperAgent\paper\{topic}\markdown/{task_id}.md")
       paper_content = re.sub(pattern1, '', paper_content, flags=re.DOTALL)
       paper_content = re.sub(pattern2, '', paper_content, flags=re.DOTALL)
    # print(paper_content)

    system_prompt="""# Task Definition
Please compress the content of the provided academic paper according to the following requirements:
1.The Methods and Experiments sections should be described in extreme detail, as they are the most crucial parts of the compression. Ensure no key technical details, experimental steps, or relevant formulas are omitted.
2.While compressing, ensure the scientific accuracy and integrity of the information are maintained.
3.Avoid redundant or repetitive content, focusing on extracting the most critical points.
4.Ensure that the overall logic and flow of the paper remain clear and coherent.
5.Ignore the Abstract Do not include or mention any part of the abstract section of the paper.

# Output Requirements
## Introduction
Provide an overview of the research background, purpose, and significance.
Highlight the research motivation, main research questions, and objectives.
Extract the most critical points, avoiding lengthy background information.

## Methods
This section should be extremely detailed.
Summarize the methodology used in the research.
Describe the research design, data sources, sample selection, experimental setup, and technical details, ensuring the methodology section remains complete.
Retain all relevant formulas, algorithms, and mathematical expressions to maintain scientific rigor.
If certain details, like sample selection or data sources, are not explicitly mentioned in the paper, feel free to skip them rather than forcing their inclusion.

## Experiments/Results
This section should be extremely detailed.
Provide a highly condensed summary of the experimental results, focusing on the main findings and their alignment (or discrepancies) with the hypotheses.
Preserve the essential elements of the experimental design, including the purpose of the experiments, hypotheses, or research questions.
Only retain the key experimental steps and procedures, while simplifying secondary details.
Ensure all critical data and results are kept intact to maintain the accuracy and completeness of the data.
Clearly indicate the baselines used in the paper.

## Conclusions
Distill the main conclusions of the study, emphasizing the research findings' contributions to the field and how they address the research questions.
Highlight the novelty of the research and its implications for future work.

## Future Work
Indicate possible directions for future research based on the current study’s results.
Discuss the limitations of the current research and provide suggestions for improvement."""

    #信息压缩：use old
    # compression_result=call_with_deep(question=f"The content is '''{paper_content}'''",system_prompt=system_prompt)

    # #use r1
    compression_result = call_with_DeepSeek_R1_250120(question=f"The content is '''{paper_content}'''", system_prompt=system_prompt)

    directory = fr"E:\PaperAgent\paper\{topic}\compression"
    if not os.path.exists(directory):
        os.makedirs(directory)

    with open(fr"{directory}/{task_id}.md", 'w', encoding='utf-8') as f:
        f.write(compression_result)

    # with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\compression/{task_id}.md",'w',encoding='utf-8') as f:
    #     f.write(compression_result)

    return compression_result

def search_releated_paper(topic,max_paper_num=1,compression=True):

    keynum=0
    relatedPaper = []
    Entities = []

    # papers=search_title_DataAndCite(query=topic,max_results=max_paper_num)
    papers = search_paper(keywords=[topic], limit=max_paper_num)

    print(papers)

    for paper in papers:
        if compression==True:
            try:
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} \033[0m")
                compression_result=information_compression(doi=paper["doi"],title=paper["title"],topic=topic)
                print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
            except:
                print(f"\033[1;31m | ERRO     | compressed paper information: {paper['title']} State: Miss! \033[0m")
                compression_result="None"
        else: compression_result="None"

        try:
            relatedPaper.append({
                "title":paper["title"],
                "abstract":paper["abstract"],
                "compression_result": compression_result
            })

            for keyword in paper["keyword"]:
                Entities.append(keyword)
        except:
            pass

    extract_entity_prompt=f"""The content is: {Entities}."""

    system_prompt = '''Please extract useful and the smallest separable keyword from the following content and return it in JSON format.
EXAMPLE JSON OUTPUT:
{
    "keywords": [
        "entity1_name",
        "entity2_name",
    ...
    ]
}'''

    print(extract_entity_prompt)

    Keywords= call_with_deep_jsonout(system_prompt=system_prompt,question=extract_entity_prompt)['keywords']

    keyword_str=""

    print(f"\033[1;32m | INFO     | Analyzing and processing Keywords:\n{Keywords}\n \033[0m")

    for keyword in Keywords:
        keynum+=1
        temp=get_description(search(query=keyword))
        if not temp:
            print(f"\033[1;31m | Warning     | {keyword}'s description  is empty \033[0m")
            keyword_str+=f"{keyword};\n"
        else:
            keyword_str+=f"{keyword}:{temp[0]};\n"
    #可能keyword解释会有问题
    print(f"\033[1;32m | INFO     | Analyzing and processing Keywords' Result:\n{keyword_str}\n \033[0m")

    print(relatedPaper)


    return keynum,relatedPaper,keyword_str

def search_releated_abstract(topic,max_paper_num=10):

    print(f"\033[1;32m | INFO     | Getting abstract paper information... \033[0m")
    # papers = search_title_DataAndCite(query=topic, max_results=max_paper_num)
    papers = search_paper(keywords=[topic], limit=max_paper_num)
    print(f"\033[1;32m | INFO     | Getting abstract paper information:OK! \033[0m")

    return papers

def extract_message(file,split_section):

    print(f"\033[1;32m | INFO     | extracting message... \033[0m")
    text=read_markdown_file(file)

    if split_section!="":
        match = re.search(fr'### {split_section}:(.*?)(?=###|\Z)', text, re.DOTALL)
        if match:
            problem_statement = match.group(1).strip()
            print(problem_statement)
    return text,problem_statement

def extract_message_review(file,split_section):

    print(f"\033[1;32m | INFO     | Info: extracting message review... \033[0m")
    text=read_markdown_file(file)

    if split_section!="":
        match = re.search(fr'# {split_section}(.*?)(?=#|\Z|\n\n)', text, re.DOTALL)
        if match:
            problem_statement = match.group(1).strip().split('\n')
            # print(problem_statement)
        else:
            print(f"\033[1;31m | ERRO     | extracting message review：erro！\033[0m")
    print(problem_statement)
    return text,problem_statement

def extract_message_review_moa(file,split_section):

    print(f"\033[1;32m | INFO     | Info: extracting message review... \033[0m")
    text=read_markdown_file(file)

    if split_section!="":
        match = re.search(fr'# {split_section}(.*?)(?=#|\Z)', text, re.DOTALL)
        # print(match)
        if match:
            problem_statement = match.group(1).strip().split('\n')
            print(problem_statement)
        else:
            print(f"\033[1;31m | ERRO     | extracting message review：erro！\033[0m")

    return text,problem_statement

def extract_technical_entities(file,split_section):

    print(f"\033[1;32m | INFO     | extracting technical entities... \033[0m")

    text,problem_statement=extract_message(file,split_section)

    system_prompt = '''Please extract key technical entities from the following text. These entities may include technical terms, tools, frameworks, algorithms, or other relevant concepts. Rank these entities based on their relevance and importance to the content, and return them in JSON format. The ranking should consider the following factors:
- Frequency of mention in the text.
- Significance in the context (e.g., whether the entity is central to the discussion or peripheral).
- Position in the text (e.g., whether it appears in the introduction, conclusion, or throughout).

EXAMPLE JSON OUTPUT:
{
"keywords": [
{
"entity": "entity1_name",
"importance_score": 0.8
},
{
"entity": "entity2_name",
"importance_score": 0.6
},
...
]
}
'''

    # extract_technical_entities：
    Keywords = call_with_deep_jsonout(system_prompt=system_prompt,question=f'The content is: {problem_statement}')['keywords']

    # # 按重要性分数降序排序，如果重要性分数相同，则按相关性分数降序排序
    # sorted_entities = sorted(Keywords , key=lambda x: (-x['importance_score']))

    sorted_entities=SearchKeyWordScore(Keywords)


    # # 输出排序后的实体列表
    # for entity in sorted_entities:
    #     print(entity)

    print(f"\033[1;32m | INFO     | extracting technical entities:OK!\n{sorted_entities} \033[0m")

    return sorted_entities,text

def extract_hypothesis(file,split_section="Hypothesis"):
    print(f"\033[1;32m | INFO     | extract hypothesis... \033[0m")

    text = read_markdown_file(file)

    # 正则表达式匹配 Hypothesis 后的内容
    pattern = re.compile(fr"{split_section} \d+:(.+?)\n", re.DOTALL)

    # 查找所有匹配项
    matches = pattern.findall(text)

    # 去除每个匹配项前后的空白字符
    hypotheses = [match.strip() for match in matches]

    # # 打印结果
    # for i, hypothesis in enumerate(hypotheses, start=1):
    #     print(f"Hypothesis {i}:\n{hypothesis}\n")

    return hypotheses

def idea_first_generate(num,topic,compression=True):

    print(f"\033[1;32m | INFO     | start generate inital idea draft... \033[0m")

    #step 1
    file_path=hypothesis_generate(topic=topic,shuffle=True,random_num=5,paper_num=10)
    # file_path=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/fromfact_{topic}_result.md"    #test use

    # step 2
    hypotheses=extract_hypothesis(file=file_path, split_section="Hypothesis") #getting Hypothesis

    hypotheses_index=0
    hypotheses_prompt=""
    for hypothesis in hypotheses:
        # print(f"Hypothesis {hypotheses_index}:{hypothesis}\n")
        hypotheses_index += 1
        hypotheses_prompt += f"\nHypothesis {hypotheses_index}: {hypothesis}"


    print(f"\033[1;32m | INFO     | The hypoyjese prompt is：{hypotheses_prompt} \033[0m")
    # print(hypotheses_prompt)

    # step 3
    keynum, relatedPaper, keyword_str = search_releated_paper(topic=topic, max_paper_num=num,compression=compression)

    title_abstract_prompt = ""

    if compression==True:
        for index,paper in enumerate(relatedPaper):
            title_abstract_prompt+=f"""\n# The {index+1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
    else:
        for index, paper in enumerate(relatedPaper):
            title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = f"""# Task Definition\nYou are going to generate a research draft that should be original, clear, feasible, relevant, and significant to astronomy field. This will be based on {hypotheses_index} related hypotheses, the title and abstract of those of {index+1} related papers in the existing literature, and {keynum} entities potentially connected to the research area.\n
Understanding of the related hypotheses, papers, and entities is essential:
- The related hypotheses should serve as a foundation or starting point for your research problem, acknowledging that they may contain flaws or inaccuracies that can be refined. They might offer insights, reveal gaps, or present contradictions that you can build upon, address, or improve upon in your problem statement.
- The related papers indicate their direct relevance and connection to the primary research topic you are focusing on, and providing additional context and insights that are essential for understanding and expanding upon the paper.
- The entities can include topics, keywords, individuals, events, or any subjects with possible direct or indirect connections to the related studies, serving as auxiliary sources of inspiration or information that may be instrumental in formulating the research problem.

I will provide the related hypotheses, papers, and entities, as follows:
# Hypotheses
{hypotheses_prompt}
{title_abstract_prompt}
# Entities: 
'''{keyword_str}'''\n

# Output Requirements
With the provided related hypotheses, papers, and entities, your objective now is to formulate a research draft that not only builds upon these existing studies but also strives to be original, clear, feasible, relevant, and significant. In addition, you also need to provide a detailed Rationale、Necessary technical details、possible datasets and Reference to ultimately form the paper title,abstract,methods and experiments. Note: The final title should be appealing, and the experimental content should include relevant baselines as much as possible.

# Respond in the following format:
### Problem:
### Rationale:
### Necessary technical details:
### Datasets:
### Paper title:
### Paper abstract:
### Methods:
### Experiments:
### Reference: 1.[Title 1], 2.[Title 2], ..., n.[Title n]"""

    # user_prompt = f"""# Task Definition\nYou are going to generate a research problem that should be original, clear, feasible, relevant, and significant to astronomy field. This will be based on {hypotheses_index} related hypotheses, the title and abstract of those of {index + 1} related papers in the existing literature, and {keynum} entities potentially connected to the research area.\n
    # Understanding of the related hypotheses, papers, and entities is essential:
    # - The related hypotheses should serve as a foundation or starting point for your research problem, acknowledging that they may contain flaws or inaccuracies that can be refined. They might offer insights, reveal gaps, or present contradictions that you can build upon, address, or improve upon in your problem statement.
    # - The related papers indicate their direct relevance and connection to the primary research topic you are focusing on, and providing additional context and insights that are essential for understanding and expanding upon the paper.
    # - The entities can include topics, keywords, individuals, events, or any subjects with possible direct or indirect connections to the related studies, serving as auxiliary sources of inspiration or information that may be instrumental in formulating the research problem.
    #
    # I will provide the related hypotheses, papers, and entities, as follows:
    # # Hypotheses
    # {hypotheses_prompt}
    # {title_abstract_prompt}
    # # Entities:
    # '''{keyword_str}'''\n
    # # Output Requirements\n
    # With the provided related papers, and entities, your objective now is to formulate a research problem that not only builds upon these existing studies but also strives to be original, clear, feasible, relevant, and significant. In addition, you also need to provide a Rationale、Necessary technical details and possible datasets to ultimately form the paper title,abstract,methods and experiments.\n
    # # Respond in the following format:
    # ### Problem:
    # ### Rationale:
    # ### Necessary technical details:
    # ### Datasets:
    # ### Paper title:
    # ### Paper abstract:
    # ### Methods:
    # ### Experiments:"""

#     user_prompt=f"""# Task Definition\nYou are going to generate a research problem that should be original, clear, feasible, relevant, and significant to astronomy field. This will be based on the title and abstract of those of {papernum} related papers in the existing literature, and {keynum} entities potentially connected to the research area.
#
# Understanding of the related papers, and entities is essential:
# - The related papers indicate their direct relevance and connection to the primary research topic you are focusing on, and providing additional context and insights that are essential for understanding and expanding upon the paper.
# - The entities can include topics, keywords, individuals, events, or any subjects with possible direct or indirect connections to the related studies, serving as auxiliary sources of inspiration or information that may be instrumental in formulating the research problem.
#
# I will provide the related papers, and entities, as follows:
# {title_abstract_prompt}
# # Entities:
# '''{keyword_str}'''
#
# # Output Requirements\n
# With the provided related papers, and entities, your objective now is to formulate a research problem that not only builds upon these existing studies but also strives to be original, clear, feasible, relevant, and significant. In addition, you also need to provide a Rationale、Necessary technical details and possible datasets  to ultimately form the paper title and abstract.
#
# Respond in the following format:
# Problem:
# Rationale:
# Necessary technical details:
# Datasets:
# paper title:
# paper abstract:"""

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_input.md",'w', encoding='utf-8') as f:
        f.write(user_prompt)

    # step 4 use old
    # first_idea_result=call_with_deep(question=user_prompt,temperature=1.5)
    #use r1
    first_idea_result = call_with_DeepSeek_R1_250120(question=user_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}.md",'w', encoding='utf-8') as f:
        f.write(first_idea_result)

    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{first_idea_result}", id_row=5, sheet_name='Sheet1')

    print(f"\033[1;32m | INFO     | the initial idea draft ok \033[0m")

    return fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}.md"

def idea_iteration_generate(num,topic,compression=True):
#     user_prompt="""You are going to generate a research problem that should be original, clear, feasible, relevant, and significant to astronomy field. This will be based on the title and abstract of the target paper, those of {len(references)} related papers in the existing literature, and {len(entities)} entities potentially connected to the research area.
#
# Understanding of the target paper, related papers, and entities is essential:
# - The target paper is the primary research study you aim to enhance or build upon through future research, serving as the central source and focus for identifying and developing the specific research problem.
# - The related papers are studies that have cited the target paper, indicating their direct relevance and connection to the primary research topic you are focusing on, and providing additional context and insights that are essential for understanding and expanding upon the target paper.
# - The entities can include topics, keywords, individuals, events, or any subjects with possible direct or indirect connections to the target paper or the related studies, serving as auxiliary sources of inspiration or information that may be instrumental in formulating the research problem.
#
# Your approach should be systematic:
# - Start by thoroughly reading the title and abstract of the target paper to understand its core focus.
# - Next, proceed to read the titles and abstracts of the related papers to gain a broader perspective and insights relevant to the primary research topic.
# - Finally, explore the entities to further broaden your perspective, drawing upon a diverse pool of inspiration and information, while keeping in mind that not all may be relevant.
#
# I am going to provide the target paper, related papers, and entities, as follows:
# Target paper title: {paper[’title’]}
# Target paper abstract: {paper[’abstract’]}
# Related paper titles: {relatedPaper[’titles’]}
# Related paper abstracts: {relatedPaper[’abstracts’]}
# Entities: {Entities}
#
# With the provided target paper, related papers, and entities, your objective now is to formulate a research problem that not only builds upon these existing studies but also strives to be original, clear, feasible, relevant, and significant. Before crafting the research problem, revisit the title and abstract of the target paper, to ensure it remains the focal point of your research problem identification process.
#
# Target paper title: {paper[’title’]}
# Target paper abstract: {paper[’abstract’]}
# Then, following your review of the above content, please proceed to generate one research
# problem with the rationale, in the format of
# Problem:
# Rationale:"""
    first_idea_path=idea_first_generate(topic=topic, num=num,compression=True)

    print(f"\033[1;32m | INFO     | start iteration 1... \033[0m")

    technical_keywords,first_idea=extract_technical_entities(first_idea_path,split_section="Paper abstract")

    text,target_paper_title=extract_message(first_idea_path,split_section="Paper title")

    print(f"\033[1;32m | INFO     | retrieve papers related to technical entities... \033[0m")

    readysearch_key=[]

    #get good keyword
    for index,keyword in enumerate(technical_keywords):
        if index > 3:
            break
        # if float(keyword['composite_score']) >= 0.9 and float(keyword['relevance_score']) >= 0.9:
        readysearch_key.append(keyword['entity'])

    readysearch_paper=search_paper(keywords=readysearch_key,limit=2) #getting technical papers

    title_abstract_prompt=""
    for index,paper in enumerate(readysearch_paper):
        try:
            state=download_pdf(url = paper['pdf'], save_dir = fr"E:\PaperAgent\paper\{topic}\pdf", file_name = sanitize_folder_name(paper['title'])+".pdf")
        except:
            print(f"\033[1;31m | ERRO     | {paper['title']}:paper download failed！！！！")

        #pdf转markdown
        if state != False:
            if compression == True:
                try:
                    print(f"\033[1;32m | INFO     | Getting compressed2 paper information: {paper['title']} \033[0m")
                    compress_paper=information_compression2(paper_pdf_path=state,topic=topic)
                    paper["compression_result"] = compress_paper
                    print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
                except:
                    print(f"\033[1;31m | ERRO     | Getting compressed paper information: {paper['title']} State: Miss! \033[0m")
                    paper["compression_result"] = "None"

            if compression == True:
                title_abstract_prompt += f"""\n# The {index+1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
            else:
                title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = f"""# Task Definition
You will receive an initial idea draft {target_paper_title}, which outlines the core problem and rationale of an innovative idea and its preliminary technical details. To further enhance the feasibility, efficiency, technical details and innovation of this idea draft, you can conduct in-depth analysis the several relevant papers provided to gain inspiration for further refining and optimizing the idea draft. Finally, you need to provide a summary of the differences between the optimized draft content and the original draft. Before optimizing the idea draft, it is essential to revisit the title and abstract to ensure they capture the reader's attention and remain the central focus of the research problem identification process.

-Technical Integration: Analyze and identify which technologies, methods, or theories from the related technical papers can be directly applied to or inspire improvements in the technical details of the idea draft.

-Performance Enhancement: Consider how advanced technologies from the papers can be leveraged to improve the performance, efficiency, or user experience of the product/service in the idea draft, including but not limited to computing speed, resource consumption, and security.

-Innovation Reinforcement: Explore whether there are novel technical perspectives or underutilized technical points in the papers that can enhance the innovation and competitiveness of the idea draft.

-Feasibility Assessment: Reassess the technical feasibility of the idea draft by combining case analyses, experimental data, or theoretical derivations from the articles. Make necessary improvements.

# The provided the related technical papers as follows:
{title_abstract_prompt}

# The initial idea draft that requires iterative optimization is as follows:
{first_idea}

# Respond in the following format:
### Problem:
### Rationale:
### Necessary technical details:
### Datasets:
### Paper title:
### Paper abstract:
### Methods:
### Experiments:
### Reference: 1.[Title 1], 2.[Title 2], ..., n.[Title n]
### Summary of the differences in this iteration:"""

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_input_iteration1.md", 'w',
              encoding='utf-8') as f:
        f.write(user_prompt)

    idea_iteration_result = call_with_deep(question=user_prompt,temperature=1.5)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_iteration1.md", 'w',
              encoding='utf-8') as f:
        f.write(idea_iteration_result)

    idea_draft=idea_iteration_result.split("### Summary of the differences in this iteration:")[0].strip()
    summary=idea_iteration_result.split("### Summary of the differences in this iteration:")[1].strip()
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{idea_draft}", id_row=6, sheet_name='Sheet1')
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{summary}", id_row=7, sheet_name='Sheet1')

    print(f"\033[1;32m | INFO     | iteration 1 ok \033[0m")

    idea_iteration_2(topic=topic,compression=True)

    # idea_iteration_3(topic=topic,compression=False)

def inspiration_generation(num,topic):

    # papers=search_releated_abstract(topic=topic,max_paper_num=num)
    keyword=[topic]
    print(keyword)
    papers=search_paper(keywords=keyword, limit=num)
    print(papers)

    title_abstract_prompt = ""
    index=0
    for index, paper in enumerate(papers):
        if 'abstract' in paper:
            index+=1
            title_abstract_prompt += f"""\n# The {index} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    system_prompt = f'''Based on the provided abstracts and titles, generate five bold and innovative hypotheses. These hypotheses should demonstrate high levels of originality, value, and innovation, while remaining reasonable, grounded, and feasible. The hypotheses can involve significant revisions to existing theories, predictions of future trends, or potential breakthroughs in specific fields. If a hypothesis relies on data or findings from a particular abstract, include a reference to that abstract using the format: 'related information[cite, paper title]'.

Requirements:

1.'cite' represents a fixed reference identifier, while 'paper title' should be replaced with the exact title of the abstract being cited.
2.All hypotheses must be original, avoid well-established conclusions, and not simply restate common knowledge.
3.Each hypothesis should have the potential for high impact, either by opening new research directions, suggesting novel industry applications, or inspiring interdisciplinary approaches.
4.Hypotheses should be grounded in the key points of one or more abstracts but not restricted to just one abstract; they should synthesize ideas where appropriate and leverage broader trends or connections across fields.
5.Ensure that the hypotheses remain grounded in logic and evidence, avoiding purely speculative, overly fanciful, or unsupported concepts.
6.Make sure that each hypothesis is concise, clearly articulated, and specific in its potential impact.'''

    user_prompt=f'''The provided titles and abstracts are as follows:
{title_abstract_prompt}

# Respond in the following format:
The hypothesis is:'''

    print(system_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/{topic}_input_hypbackground.md", 'w',
              encoding='utf-8') as f:
        f.write(user_prompt)

    result=call_with_deep(system_prompt=system_prompt,question=user_prompt)

    print(result)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/{topic}_input_result.md", 'w',
              encoding='utf-8') as f:
        f.write(result)

def remove_number_prefix(paragraph):
    # # 定义一个正则表达式模式，用于匹配句子开头的数字和随后的句点空格
    # pattern = r'^\d+\. '
    # # 利用re.sub函数，将匹配到的部分替换为空字符串，以此移除它
    # modified_sentence = re.sub(pattern, '', sentence)
    # return modified_sentence
    # 定义一个正则表达式模式，用于匹配句子开头的数字和随后的句点空格
    pattern = r'^\d+\. |(?<=\n)\d+\. '
    # 利用re.sub函数，将匹配到的部分替换为空字符串，以此移除它
    modified_paragraph = re.sub(pattern, '', paragraph, flags=re.MULTILINE)
    return modified_paragraph

def get_related_keyword(topic):

    print(f"\033[1;32m | INFO     | geting related keyword... \033[0m")

    user_prompt = f"""# Task Definition
Based on the research topic '{topic}', recommend five relevant technical keywords or solution approaches that are specific, actionable, and directly related to innovative techniques or methods in this area.

# Output Format:
["Keyword 1", "Keyword 2", "Keyword 3", "Keyword 4", "Keyword 5"]"""

    #use old
    # result = call_with_deep(system_prompt="You are a helpful assistant.",question=user_prompt)
    #usl r1
    result = call_with_DeepSeek_R1_250120(system_prompt="You are a helpful assistant.", question=user_prompt)

    print(f"\033[1;32m | INFO     | The related keyword is :{result} \033[0m")

    print(f"\033[1;32m | INFO     | geting related keyword:OK! \033[0m")

    return ast.literal_eval(result)

def fact_information_extraction(num,topics,maintopic):

    print(f"\033[1;32m | INFO     | fact information extraction... \033[0m")

    # 加载现有的 XLSX 文件
    file_path = fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\fact/fact_information_{maintopic}.xlsx"  # 替换为你的文件路径
    wb = Workbook()
    ws = wb.active
    start_row = ws.max_row + 1  # 在最后一行之后开始追加

    keyword=topics

    # fact_information_extraction
    papers=search_paper(keywords=keyword, limit=num)
    # print(papers)

    system_prompt = f"""# Task Definition: Please extract the objective factual information from the following provided paper title and abstract, and output it. Ensure that the extracted information meets the following requirements:

1. **Avoid Demonstrative Pronouns**: Do not use demonstrative pronouns such as 'this', 'that', 'these', 'those', 'this work', 'this survey', 'the study', etc. Instead, use 'Related research'. Ensure that each piece of information is self-contained and clear.
2. **Dataset Description Standards**: When describing datasets, not only specify which dataset was used, but also clearly state the source of the dataset and its intended purpose (e.g., The data comes from [data source] can be used in [research topic]. ).
3. **Result Presentation Standards**: For experimental results, avoid using ambiguous phrases like "the results indicate" and instead use "Related research show [experimental results]," ensuring that the content of the experimental results is clear and specific.
4. **Accuracy**: All information should be an explicit statement from the abstract, without personal speculation or interpretation.
5. **Completeness**: Include all main factual points from the abstract, avoiding the omission of key information.
6. **Conciseness**: Each fact should be a concise and clear statement.
7. **Structured Format**: Present each factual information in a numbered list format, with each fact on a separate line.
8. **Independence and Clarity**: Each piece of output information must be able to stand alone without context, with clear meaning, free from ambiguity or unclear references.

# Example output format:

1.Related research have shown that [research topic].
2.Related research have used [methods], and find that [experimental results].
3.The data comes from [data source] can be used in [research topic].
4.Related research have shown that [experimental results].
5.Related research have found that [conclusion]."""

    for paper in papers:
        if 'abstract' in paper:
            user_prompt = f"""Now,please following these rules to extract the factual information from following paper:\ntitile:{paper['title']}\nabstract:\n{paper['abstract']}\n"""

            #use old
            # result = call_with_deep(system_prompt=system_prompt, question=user_prompt)

            #use r1
            result = call_with_DeepSeek_R1_250120(system_prompt=system_prompt, question=user_prompt)

            # 追加数据到 Excel 中
            ws.cell(row=start_row, column=1, value=paper['topic'])
            ws.cell(row=start_row, column=2, value=paper['title'])
            ws.cell(row=start_row, column=3, value=paper['abstract'])
            ws.cell(row=start_row, column=4, value=result)
            start_row += 1  # 移动到下一行

    # 保存 Workbook 到文件（会覆盖原文件，但数据是在原文件基础上追加的）
    wb.save(file_path)

    print(f"\033[1;32m | INFO     | fact information extraction:ok! \033[0m")

    return file_path

def hypothesis_generate(topic,shuffle=False,random_num=5,paper_num=10):

    topics=get_related_keyword(topic=topic) #other maybe related keywords

    file_path=fact_information_extraction(num=paper_num, topics=topics,maintopic=topic) #getting related paper abstract

    print(f"\033[1;32m | INFO     | hypothesis_generate... \033[0m")

    wb = load_workbook(filename=file_path)
    ws = wb.active

    fact_information=[]

    for keyword in topics:
        # print(keyword)
        index=0
        for paper in ws.iter_rows(min_row=2,values_only=False):
            if paper[0].value==keyword and index<random_num:
                # print(index)
                temp=remove_number_prefix(paper[3].value)
                # print(temp)
                fact_information += temp.split('\n')
                index+=1

    if shuffle==True:
        random.shuffle(fact_information)

    Known_Information=""
    index=0
    for information in fact_information:
        if information!="" and information!="\n" :
            Known_Information+=f"{index+1}. {information}\n"
            index+=1

    print(f"\033[1;32m | INFO     | the fact information is :\n{Known_Information} \033[0m")

    user_prompt=f"""# Task Definition: Based on  the target research topic '{topic}' and the following known information, derive 5 new hypotheses by combining the theories of inductive and deductive reasoning

1. Theory of Inductive Reasoning: Inductive reasoning is the process of summarizing general laws or hypotheses from specific instances. Based on the known information, we can use inductive reasoning to derive some new hypotheses.
2. Theory of Deductive Reasoning: Deductive reasoning is a logical reasoning method where conclusions are necessarily drawn from one or more premises.

# Known Information:
{Known_Information}

# Requirements:

1. Cite relevant theories as the basis for derivation.
2. Gradually analyze the relationships between the known information and apply relevant theories for reasoning.
3. Propose reasonable, novel, high-quality, and worthwhile new hypotheses, and briefly explain the derivation process and theoretical basis.
4. Skip over potentially low-quality or useless known information.
5. Assume the content should be closely aligned with target research topic.

# Output Format:
Hypothesis [Number]: [New Hypothesis]

Derivation Process:
- Information Considered: [Information 1], [Information 2], ..., [Information n]
- Theory Applied: [Relevant Theory]
- Reasoning: Based on the information and theoretical foundation, it is derived that [New Hypothesis].

Theoretical Basis:
- Theory: [Brief explanation of relevant theory].
"""

    #use old
    # result = call_with_deep(system_prompt="You are a research expert.", question=user_prompt,temperature=1.5)

    #use r1
    result = call_with_DeepSeek_R1_250120(system_prompt="You are a research expert.", question=user_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/fromfact_{topic}_input.md",'w', encoding='utf-8') as f:
        f.write(user_prompt)
    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/fromfact_{topic}_result.md",'w', encoding='utf-8') as f:
        f.write(result)

    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{Known_Information}", id_row=3, sheet_name='Sheet1')
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{result}", id_row=4, sheet_name='Sheet1')

    print(f"\033[1;32m | INFO     | hypothesis_generate:ok! \033[0m")

    return fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/fromfact_{topic}_result.md"

def review_mechanism(topic,idea_draft=""):
    # system_prompt = "You are a helpful assistant."
    print(f"\033[1;32m | INFO     | now review is running... \033[0m")

    system_prompt = f"""# Task Definition
Please review the provided research idea draft with the rigor and critical eye of a research reviewer, including suggestions for improvement. The output should include the following:

# Analysis and Reflection:
1. **Insufficient Detail**: Identify any shortcomings in technical details, data selection, or methodological design within the draft.
2. **Other Areas for Improvement**: For example, aspects such as innovation, alignment with target user needs, and consistency with research trends in the relevant field.
   - **Innovative Improvements**: Suggest analyzing whether there are alternative methods, models, or technologies, and if there’s a need to further explore the unique contributions of the research.
   - **Needs Alignment**: Further analyze the alignment of the research with its target users and their needs, identifying any deviations from user requirements.
   - **Consistency with Field Trends**: Suggest evaluating whether the research aligns with current technological and methodological trends in the field or if it addresses cutting-edge issues.
3. **Scalability and Future Application Prospects**: Analyze whether the research has scalability and potential application value in the future.
4. **Self-Reflection**: Suggest angles for researchers to engage in self-reflection (e.g., the significance of the research question, the existence of alternative methods, depth of understanding existing work, etc.).
   - Recommend increasing the "in-depth exploration of related research and data" to confirm comprehensive understanding and reasonable utilization of existing studies.
   - "Reasonableness of Research Hypothesis" can be a key point of self-reflection, i.e., whether the research hypothesis aligns with the current context.
   - "Feasibility Reflection of the Research" can also help consider the technical difficulty and resource investment of the project, determining if the necessary conditions for completion are present.

# Next Steps for Optimization:
Clearly define improvement objectives based on the issues identified in the draft. Provide preliminary directions or plans for each objective.

# Iterative Optimization Search Keywords:
Key technologies and themes for subsequent literature retrieval. Provide search suggestions for each keyword (e.g., relevant fields).

# A reference format for orderly presentation of the review and analysis:
- # Analysis and Reflection**:
  - **Insufficient Detail**:
    - [List specific details that are lacking]
  - **Other Areas for Improvement**:
    - Innovative Improvements: [List suggestions]
    - Needs Alignment: [List alignment status]
    - Consistency with Field Trends: [List consistency analysis]
    - Scalability and Future Application Prospects: [List analysis]
  - **Self-Reflection**:
    - In-depth Exploration of Related Research and Data: [List reflection points]
    - Reasonableness of Research Hypothesis: [List reflection points]
    - Feasibility Reflection of the Research: [List reflection points]

- # Next Steps for Optimization**:
  - [Objective 1]
  - [Objective 2]
  - ...

- # Iterative Optimization Search Keywords**:
  - [Keyword 1] - [Search suggestion]
  - [Keyword 2] - [Search suggestion]
  - ..."""

    user_prompt=f"""# Idea Draft\n{idea_draft}"""

    result = call_with_deep(system_prompt=system_prompt, question=user_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\review/{topic}_review.md",'w', encoding='utf-8') as f:
        f.write(result)

    text,optimize_messages=extract_message_review(file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\review/{topic}_review.md", split_section="Iterative Optimization Search Keywords")

    keywords=[]

    for optimize_message in optimize_messages:
        temp=optimize_message.split(' - ')
        # print(temp)
        keywords.append({'keyword':temp[0],'describe':temp[1]})

    print(keywords)

    print(f"\033[1;32m | INFO     | now review is running:OK! \033[0m")

    return keywords

def idea_iteration_2(topic,compression = True):
    print(f"\033[1;32m | INFO     |  start idea draft iter 2... \033[0m")

    idea_draft, target_paper_title = extract_message(file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_iteration1.md",
                                                     split_section="Paper title") #getting initial draft title

    information=review_mechanism(topic=topic,idea_draft=idea_draft)

    text, next_optimization = extract_message_review(
        file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\review/{topic}_review.md",
        split_section="Next Steps for Optimization") #getting review

    target_next_optimization=""
    for opt in next_optimization:
        target_next_optimization+=f"\n{opt}"

    print(f"\033[1;32m | INFO     |  next optimization target is:\n {target_next_optimization} \033[0m")

    print(target_next_optimization)

    readysearch_key=[]
    optimization_keywords=""

    for keyword in information:
        match = re.search(fr'\*\*(.*)\*\*', keyword['keyword']).group(1).strip()
        readysearch_key.append(match)
        optimization_keywords+=f"\n{keyword['keyword']}:{keyword['describe']}"

    print(readysearch_key)

    readysearch_paper=search_paper(keywords=readysearch_key,limit=2)

    # print(readysearch_paper)

    title_abstract_prompt=""
    for index,paper in enumerate(readysearch_paper):
        try:
            state=download_pdf(url = paper['pdf'], save_dir = fr"E:\PaperAgent\paper\{topic}\pdf", file_name = sanitize_folder_name(paper['title'])+".pdf")
        except:
            print(f"\033[1;31m | ERRO     | {paper['title']}:paper download failed！！！！")

        #pdf转markdown
        if state != False:
            if compression == True:
                try:
                    print(f"\033[1;32m | INFO     | Getting compressed2 paper information: {paper['title']} \033[0m")
                    compress_paper=information_compression2(paper_pdf_path=state,topic=topic)
                    paper["compression_result"] = compress_paper
                    print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
                except:
                    print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: Miss! \033[0m")
                    paper["compression_result"] = "None"
            if compression == True:
                title_abstract_prompt += f"""\n# The {index+1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
            else:
                title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = f"""# Task Definition
You will receive an idea draft, which outlines the core problem and rationale of an innovative idea and its preliminary technical details. To further enhance the feasibility, efficiency, or innovativeness of this idea, you should first refer to the provided "# Next Steps for Optimization" and "# Optimization Keywords" content, which will provide you with specific directions for optimization. Subsequently, you can conduct in-depth analysis the several relevant papers provided to gain inspiration for further refining and optimizing the idea draft. Finally, you need to provide a summary of the differences between the optimized draft content and the original draft. Before optimizing the idea draft, it is essential to revisit the title and abstract to ensure they capture the reader's attention and remain the central focus of the research problem identification process.

# Next Steps for Optimization:
{target_next_optimization}

# Optimization Keywords：
{optimization_keywords}

# The provided the related papers as follows:
{title_abstract_prompt}

# The idea draft that requires iterative optimization is as follows:
{idea_draft.split("### Summary of the differences in this iteration:")[0].strip()}

# Respond in the following format:
### Problem:
### Rationale:
### Necessary technical details:
### Datasets:
### Paper title:
### Paper abstract:
### Methods:
### Experiments:
### Reference: 1.[Title 1], 2.[Title 2], ..., n.[Title n]
### Summary of the differences in this iteration:"""

    # print(user_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_input_iteration2.md", 'w',
              encoding='utf-8') as f:
        f.write(user_prompt)

    # idea_iteration_result = call_with_deep(question=user_prompt)
    # idea_iteration_result = call_with_deep(system_prompt="You are a research assistant.",question=user_prompt,temperature=1.5)  #原始版本迭代
    idea_iteration_result = moa_idea_iteration(topic=topic, user_prompt=user_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_iteration2.md", 'w',
              encoding='utf-8') as f:
        f.write(idea_iteration_result)

    idea_draft=idea_iteration_result.split("### Summary of the differences in this iteration:")[0].strip()
    summary=idea_iteration_result.split("### Summary of the differences in this iteration:")[1].strip()
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{idea_draft}", id_row=8, sheet_name='Sheet1')
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{summary}", id_row=9, sheet_name='Sheet1')

    print(f"\033[1;32m | INFO     | iteration 2 ok \033[0m")

def idea_iteration_3(topic,compression = False):
    print(f"\033[1;32m | INFO     |  start idea draft iter 3... \033[0m")

    idea_draft, target_paper_title = extract_message(file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_iteration2.md",
                                                     split_section="Paper title")

    print(idea_draft.split("### Summary of the differences in this iteration:")[0].strip())
    moa_table(topic=topic,draft=idea_draft.split("### Summary of the differences in this iteration:")[0].strip())

    # print(target_paper_title)

    _, next_optimization = extract_message_review_moa(
        file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\moa/{topic}_review_moa.md",
        split_section="Overall Opinions:")

    text,optimize_messages=extract_message_review(
        file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\moa/{topic}_review_moa.md",
        split_section="Iterative Optimization Search Keywords:")

    keywords=[]

    for optimize_message in optimize_messages:
        temp=optimize_message.split(' - ')
        # print(temp)
        keywords.append({'keyword':temp[0],'describe':temp[1]})

    target_next_optimization=""
    for opt in next_optimization:
        target_next_optimization+=f"\n{opt}"

    print(f"\033[1;32m | INFO     |  next optimization target is :\n{target_next_optimization} \033[0m")

    readysearch_key=[]
    optimization_keywords=""

    for keyword in keywords:
        match = re.search(fr'\*\*(.*)\*\*', keyword['keyword']).group(1).strip()
        readysearch_key.append(match)
        optimization_keywords+=f"\n{keyword['keyword']}:{keyword['describe']}"

    print(f"\033[1;32m | INFO     |  ready search key :\n{readysearch_key} \033[0m")

    readysearch_paper=search_paper(keywords=readysearch_key,limit=2)

    print(readysearch_paper)

    title_abstract_prompt=""
    for index,paper in enumerate(readysearch_paper):
        try:
            state=download_pdf(url = paper['pdf'], save_dir = fr"E:\PaperAgent\paper\{topic}\pdf", file_name = sanitize_folder_name(paper['title'])+".pdf")
        except:
            print(f"\033[1;31m | ERRO     | {paper['title']}:paper download failed！！！！")

        #pdf转markdown
        if state != False:
            if compression == True:
                try:
                    print(f"\033[1;32m | INFO     | Getting compressed2 paper information: {paper['title']} \033[0m")
                    compress_paper=information_compression2(paper_pdf_path=state,topic=topic)
                    paper["compression_result"] = compress_paper
                    print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: OK! \033[0m")
                except:
                    print(f"\033[1;32m | INFO     | Getting compressed paper information: {paper['title']} State: Miss! \033[0m")
                    paper["compression_result"] = "None"
            if compression == True:
                title_abstract_prompt += f"""\n# The {index+1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n## content\n{paper['compression_result']}\n"""
            else:
                title_abstract_prompt += f"""\n# The {index + 1} related paper\n## title\n{paper['title']}\n## abstract\n{paper['abstract']}\n"""

    user_prompt = f"""# Task Definition
You will receive an idea draft, which outlines the core problem and rationale of an innovative idea and its preliminary technical details. To further enhance the feasibility, efficiency, or innovativeness of this idea, you should first refer to the provided "# Objectives for optimizing the next idea draft" and "# Keywords for optimizing the next idea draft" content, which will provide you with specific directions for optimization. Subsequently, you can conduct in-depth analysis the several relevant papers provided to gain inspiration for further refining and optimizing the idea draft. Finally, you need to provide a summary of the differences between the optimized draft content and the original draft. Before optimizing the idea draft, it is essential to revisit the title and abstract to ensure they capture the reader's attention and remain the central focus of the research problem identification process.

# Objectives for optimizing the next idea draft:
{target_next_optimization}

# Keywords for optimizing the next idea draft：
{optimization_keywords}

# The provided the related papers as follows:
{title_abstract_prompt}

# The idea draft that requires iterative optimization is as follows:
{idea_draft.split("### Summary of the differences in this iteration:")[0].strip()}

# Respond in the following format:
### Problem:
### Rationale:
### Necessary technical details:
### Datasets:
### Paper title:
### Paper abstract:
### Methods:
### Experiments:
### Reference: 1.[Title 1], 2.[Title 2], ..., n.[Title n]
### Summary of the differences in this iteration:"""

    # print(user_prompt)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_input_iteration3.md", 'w',
              encoding='utf-8') as f:
        f.write(user_prompt)

    # idea_iteration_result = call_with_deep(question=user_prompt)
    # idea_iteration_result = call_with_deep(system_prompt="You are a research assistant.",question=user_prompt,temperature=1.5)

    idea_iteration_result =call_with_qwenplus(system_prompt="You are a research assistant.",question=user_prompt,temperature=1.5)

    with open(fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/{topic}_iteration3.md", 'w',
              encoding='utf-8') as f:
        f.write(idea_iteration_result)

    idea_draft=idea_iteration_result.split("### Summary of the differences in this iteration:")[0].strip()
    summary=idea_iteration_result.split("### Summary of the differences in this iteration:")[1].strip()
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{idea_draft}", id_row=10, sheet_name='Sheet1')
    update_excel_row(file_path=save_exp, target_topic=topic, new_data=f"{summary}", id_row=11, sheet_name='Sheet1')

    print(f"\033[1;32m | INFO     | iteration 3 ok \033[0m")

# def all_process():


if __name__ =='__main__':
    # search_releated_paper()
    idea_first_generate(topic="pulsar candidate classification",num=5)
    # pdf2md_mineruapi(file_path=r"E:\PaperAgent\paper\Pulsar Candidate Classification\pdf/test.pdf", topic="Pulsar Candidate Classification")
    # idea_iteration_generate(topic="double degenerate supernova", num=5)

    # erro=[]
    # keys=["Automated Redshift Measurement","Black Hole Detection and Mass Estimation","Detection of Star-Forming Regions","Star Cluster Evolution Modeling","Quasar Light Curve Analysis","Automated Gravitational Lensing Detection","Astronomical Data Anomaly Detection","Solar Activity Prediction","AI-Assisted Astronomical Instrument Calibration","Optimization of Stellar Evolution Models","Light Curve","Space Weather Forecasting","Gravitational Wave Source Classification","Study of the Milky Way's Central Black Hole","Galaxy Merger Event Detection","AI-Generated Star Maps","Exoplanet Atmosphere Composition Analysis","Time-Series Astronomy","Detection of Stellar Ejections","AI Automated Telescope Control","Estimating Cosmic Expansion Rate","Study of Dark Energy Properties","Supernova Remnants","CV eruption","Early light curve of TDE","Narrow band photometric","Shock breakthrough","Red supergiant problem","Circumstellar dust","Simple degenerate","bidegenerate","Binary precursor star","Massive stellar material loss","flare star","Compact object","Compact binaries","Stellar rotation","stellar cycle","starquake","Be star eruption","exoplanets detection","Center engine","Hubble crisis","Nebular spectrum","Solar Flare Prediction","Embodied Intelligence Telescope","The AlphaFold Moment for Astronomical foundation models"]
    # for key in keys:
    #     try:
    #         LLM_API.total_tokens_used=0
    #         idea_iteration_generate(topic=key,num=5)
    #         print(f"Over:{key}")
    #     except:
    #         LLM_API.total_tokens_used = 0
    #         print(f"!!!!!!!!!!!!!!!!!!!!!!!!!!!!erro:{key}")
    #         erro.append(key)
    # print(erro)

    # idea_iteration_3(topic="Astronomical foundation models")

    # idea_iteration_result = call_with_qwenplus(system_prompt="You are a research assistant.", question=user,
    #                                            temperature=1.5)

    # _, next_optimization = extract_message_review_moa(
    #     file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\moa/photometric classification (stars and calaxies,stars and quasars,supernovas)_review_moa.md",
    #     split_section="Overall Opinions:")
    # print(next_optimization)
    # idea_iteration_3(topic="photometric classification (stars and calaxies,stars and quasars,supernovas)")
    # text,optimize_messages=extract_message_review(
    #     file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\moa/photometric classification (stars and calaxies,stars and quasars,supernovas)_review_moa.md",
    #     split_section="Iterative Optimization Search Keywords:")

    # print(optimize_messages)
    # keywords=[]
    #
    # for optimize_message in optimize_messages:
    #     temp=optimize_message.split(' - ')
    #     match = re.search(r'\*\*(.*?)\*\*', temp[0])
    #     # print(temp)
    #     if match:
    #         keywords.append({'keyword': match.group(1), 'describe': temp[1]})
    #     else:
    #         keywords.append({'keyword': temp[0], 'describe': temp[1]})
    #
    # print(keywords)
    # extract_message_review_moa(file=r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\review/photometric classification (stars and calaxies,stars and quasars,supernovas)_review_moa.md",
    #                            split_section='')

    # idea_draft, target_paper_title = extract_message(file=fr"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/photometric classification (stars and calaxies,stars and quasars,supernovas)_iteration2.md",
    #                                                  split_section="Paper title")
    #
    # moa_table(topic='photometric classification (stars and calaxies,stars and quasars,supernovas)',draft=idea_draft)
    #
    # test=read_markdown_file(file_path=r'C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/Pulsar Candidate Classification_input_iteration3.md')
    # moa_table(topic='photometric classification (stars and calaxies,stars and quasars,supernovas)', draft=test)
    # idea_iteration_result = call_with_qwenplus(system_prompt="You are a research assistant.", question=test,
    #                                        temperature=1.5)
    # #
    # print(idea_iteration_result)

    # moa_idea_iteration(topic="Pulsar Candidate Classification", user_prompt=test)
    # information_compression(doi="10.48550/arXiv.2410.13650", title="Millisecond pulsars phenomenology under the light of graph theory",topic="Supernova Explosion Prediction")
    # review_mechanism(topic="Pulsar Candidate Classification")
    # idea_iteration_3(topic="photometric classification (stars and calaxies,stars and quasars,supernovas)")
    # technical_keywords,first_idea=extract_technical_entities(file=r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/Pulsar CNN.md", split_section="Paper abstract")
    #
    # print(technical_keywords)

    # text,paper_title=extract_message(r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\idea/Pulsar CNN.md", split_section="Paper title")
    # print(paper_title)

    # inspiration_generation(num=10, topic="Pulsar CNN")

    # get_related_keyword(topic="Pulsar Candidate Classification")
    #
    # fact_information_extraction(num=10, topic=["pulsar candidate classification","Machine Learning Algorithms", "Feature Engineering", "Deep Learning Models", "Signal Processing Techniques", "Ensemble Methods"])

    # hypothesis_generate(topic="Pulsar Candidate Classification",shuffle=True,random_num=5,paper_num=10)

    # extract_hypothesis(file=r"C:\Users\10412\Desktop\多模态大语言模型\Code\天文Code\PaperAgent\Paper_Ori\hypotheses/fromfact_Pulsar Candidate Classification_result.md")